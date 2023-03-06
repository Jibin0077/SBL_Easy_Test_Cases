import os

def current_diectory():
    relpath = os.getcwd()
    # print(os.getcwd())
    curr_dict = relpath
    return curr_dict

# print(current_diectory())

def lib_append_to_Data_list(Ext_era,Ext_sur_name,Ext_fore_name,Ext_rank,Ext_service_number,Ext_decoration,Ext_birth_place,Ext_death_place,Ext_theatre_death,Ext_death_cause,Ext_roll,Ext_unit_name,Ext_other_detail,Ext_Record_url):

    list_details=[]
    #if Code1!=None and Code2!=None:
    list_details.extend([Ext_era,Ext_sur_name,Ext_fore_name,Ext_rank,Ext_service_number,Ext_decoration,Ext_birth_place,Ext_death_place,Ext_theatre_death,Ext_death_cause,Ext_roll,Ext_unit_name,Ext_other_detail,Ext_Record_url])
    #else:

        #list_details.extend([Client,InvoiceNumber,Insurance,PTname,DOS,ServiceCode,MPI,CCN_Number,RemitDate,DEB,PaidAmount,Status,HCPCS_Code,str(FromDate),str(NumServices),SubmittedAmount,AllowedAmount])
    return list_details

# -------------------------------------------------------

# +
from csv import writer

def Write_data_Output(saveloc,List):  
    
    #header= ['Era','Surname','Forename','Rank','Service Number','Decoration','Place of Birth','Theatre of Death','Cause of Death','SNWM Roll','Unit Name','Other Detail','Record Url']
    with open(saveloc+'\InputData.csv','a',newline='') as f_object:

# Pass this file object to csv.writer()
        # and get a writer object
        writer_object = writer(f_object)
# Pass the list as an argument into
        # the writerow()
        #writer_object.writerow(header)
        writer_object.writerow(List)
#Close the file object
        f_object.close()

#--------------Updating Status in Config "Controller" sheet of Excel File------------------

def create_columns_if_not_exists(ws, columns):
    """
    Creates the columns if they do not exist.

    parameters:
        `ws` (object): Worksheet object.
        `columns` (list): List of column headers.

    Returns:
        ws (Worksheet): Worksheet object.
        colss (dict): Dictionary of column names and their respective column numbers.
    """
    colss = {}
    for column in columns:
        for col in ws.iter_cols():
            if column.lower() == col[0].value.lower():
                colss[column] = col[0].column-1
                break
        else:
            ws.insert_cols(ws.max_column+1)
            statuscell = ws.cell(row=1, column=ws.max_column+1)
            statuscell.value = column
            colss[column] = statuscell.column-1
    return ws, colss

def update_excel_value(filename, criteriacolumnheader, criteriacellvalue, statuscolumnheader, statusvalue, commentscolmheader, commentvalue):
    """
    Updates the excel file with the given values.

    Returns:
        True if the value is updated.
    """
    import openpyxl
    wb = openpyxl.load_workbook(filename)
    ws = wb['Controller']
    ws, statuscolumn = create_columns_if_not_exists(ws, [statuscolumnheader])
    ws, commentcolumn = create_columns_if_not_exists(ws, [commentscolmheader])
    for col in ws.iter_cols():
        if criteriacolumnheader == col[0].value:
            for row in ws.iter_rows():
                if criteriacellvalue == str(row[col[0].column-1].value):
                    row[statuscolumn[statuscolumnheader]].value = statusvalue
                    row[commentcolumn[commentscolmheader]].value = commentvalue
                    wb.save(filename)
                    return True
# update_excel_value("C:/Users/Q0041/Documents/Robots/RoboCorp/SBL_Easy/Config/config.xlsx","TEST CASE ID","TC_03","Status","pass","Comments","Not Found")

        

